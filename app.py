import streamlit as st
import pandas as pd
import io
import os

# إعدادات الصفحة
st.set_page_config(page_title="توزيع كراسات الإجابة", layout="wide")

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;700&display=swap');
    * { font-family: 'Tajawal', sans-serif !important; }
    .stApp, .stMarkdown, .stText, p, label, input, .stSelectbox, .stDateInput { 
        text-align: right !important; direction: rtl !important; font-size: 16px !important;
    }
    h1 { text-align: center !important; color: #1E3A8A !important; font-weight: 700 !important; }
    h3 { 
        color: #004d40 !important; font-weight: 700 !important; padding-top: 15px; 
        border-bottom: 2px solid #b6e3f4; padding-bottom: 10px; margin-bottom: 20px; 
        text-align: right !important; direction: rtl !important;
    }
    .stButton > button { font-size: 18px !important; font-weight: bold !important; width: 100% !important; }
    
    [data-testid="stDataFrame"] div[role="gridcell"], 
    [data-testid="stDataFrame"] div[role="columnheader"] {
        text-align: right !important; justify-content: flex-end !important; font-size: 15px !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# --- تنسيق الهيدر (الشعارات والعنوان) ---
col_left, col_space, col_right = st.columns([1, 3, 1])
with col_left:
    if os.path.exists("logo_unit.jpg"): st.image("logo_unit.jpg", use_container_width=True)
    elif os.path.exists("logo_unit.png"): st.image("logo_unit.png", use_container_width=True)
with col_space:
    st.markdown("<div style='display: flex; justify-content: center; align-items: center; height: 100%; margin-top: 20px;'><h1 style='margin: 0;'>توزيع كراسات الإجابة (الشجرة)</h1></div>", unsafe_allow_html=True)
with col_right:
    if os.path.exists("logo_faculty.jpg"): st.image("logo_faculty.jpg", use_container_width=True)
    elif os.path.exists("logo_faculty.png"): st.image("logo_faculty.png", use_container_width=True)
st.markdown("---")

ARABIC_LETTERS = "أ ب ج د هـ و ز ح ط ي ك ل م ن س ع ف ص ق ر ش ت ث خ ذ ض ظ غ".split(" ")
ENGLISH_LETTERS = "A B C D E F G H I J K L M N O P Q R S T U V W X Y Z".split(" ")

if 'base_df' not in st.session_state: st.session_state.base_df = None

# --- مرحلة رفع الملف ---
if st.session_state.base_df is None:
    st.markdown("<div dir='rtl' style='text-align: right; background-color: #e8f4f8; padding: 15px; border-radius: 5px; color: #004d40; border: 1px solid #b6e3f4; margin-bottom: 15px; font-size: 18px; font-weight: bold;'>ℹ️ ملاحظة: يرجى رفع ملف الإكسيل الأساسي (يحتوي على: رقم اللجنة، مكان اللجنة)</div>", unsafe_allow_html=True)
    uploaded_file = st.file_uploader("", type=['xlsx'])
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            df.columns = df.columns.str.strip()
            if "رقم اللجنة" in df.columns and "مكان اللجنة" in df.columns:
                st.session_state.base_df = df[['رقم اللجنة', 'مكان اللجنة']].copy()
                st.session_state.base_df['عدد الحضور'] = 0
                st.rerun()
            else:
                st.error("الملف يجب أن يحتوي على عمودين باسم: 'رقم اللجنة' و 'مكان اللجنة'")
        except Exception as e:
            st.error(f"حدث خطأ أثناء قراءة الملف: {e}")

# --- مرحلة واجهة العمل ---
if st.session_state.base_df is not None:
    st.markdown("<h3>البيانات الأساسية</h3>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1: 
        control_name = st.text_input("الكنترول:")
        exam_date = st.text_input("تاريخ الامتحان:")
    with col2: 
        course_name = st.text_input("المقرر:")
        lang = st.selectbox("لغة الشجرة:", ["عربي", "إنجليزي"])
        
    st.markdown("<h3>إدخال أعداد الحضور</h3>", unsafe_allow_html=True)
    input_display_cols = ['عدد الحضور', 'مكان اللجنة', 'رقم اللجنة']
    df_for_editor = st.session_state.base_df[input_display_cols]
    
    edited_df = st.data_editor(df_for_editor, disabled=["رقم اللجنة", "مكان اللجنة"], hide_index=True, use_container_width=True, height=(len(st.session_state.base_df) + 1) * 38)
    
    calc_button = st.button("حساب الشجرة وتوليد النتيجة", type="primary")

    if calc_button:
        with st.spinner('جاري الحساب...'):
            letters = ARABIC_LETTERS if lang == "عربي" else ENGLISH_LETTERS
            tree_results = []
            current_letter_idx = 0
            current_paper = 1
            
            for index, row in edited_df.iterrows():
                try: attendance = int(row['عدد الحضور'])
                except: attendance = 0
                    
                if attendance <= 0:
                    tree_results.append("")
                    continue
                
                remaining = attendance
                committee_tree = []
                while remaining > 0:
                    if current_letter_idx >= len(letters): break
                    available_in_current = 101 - current_paper
                    letter = letters[current_letter_idx]
                    if remaining <= available_in_current:
                        end_paper = current_paper + remaining - 1
                        if lang == "عربي":
                            committee_tree.append(f"{letter} {end_paper}-{current_paper}")
                        else:
                            committee_tree.append(f"{letter} {current_paper}-{end_paper}")
                        current_paper = end_paper + 1
                        remaining = 0
                        if current_paper > 100:
                            current_letter_idx += 1
                            current_paper = 1
                    else:
                        if lang == "عربي":
                            committee_tree.append(f"{letter} 100-{current_paper}")
                        else:
                            committee_tree.append(f"{letter} {current_paper}-100")
                        remaining -= available_in_current
                        current_letter_idx += 1
                        current_paper = 1
                tree_results.append(", ".join(committee_tree))
            
            result_df = edited_df.copy()
            col_name = 'الشجرة (عربي)' if lang == "عربي" else 'الشجرة (انجليزي)'
            result_df[col_name] = tree_results
            
            # --- إضافة الإجمالي ---
            total_attendance = pd.to_numeric(result_df['عدد الحضور'], errors='coerce').fillna(0).sum()
            total_row = pd.DataFrame({'رقم اللجنة': ['الإجمالي'], 'مكان اللجنة': [''], 'عدد الحضور': [int(total_attendance)], col_name: ['']})
            result_df_with_total = pd.concat([result_df, total_row], ignore_index=True)
            
            display_cols = ['رقم اللجنة', 'مكان اللجنة', 'عدد الحضور', col_name]
            result_df_display = result_df_with_total[display_cols]
            
            # --- إرجاع الجدول للموقع ---
            st.markdown(f"<div dir='rtl' style='text-align: right; color: green; font-size: 18px; font-weight: bold; margin-bottom: 15px;'>✅ تم الحساب بنجاح! إجمالي عدد الحضور: {int(total_attendance)} طالب</div>", unsafe_allow_html=True)
            output_table_height = (len(result_df_display) + 1) * 38
            styled_output = result_df_display.style.set_properties(**{'text-align': 'right', 'direction': 'rtl'})
            st.dataframe(styled_output, hide_index=True, use_container_width=True, height=output_table_height)
            
            # --- صناعة الإكسيل المطابق للصورة 100% ---
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # كتابة البيانات بداية من الصف 5 عشان نسيب مساحة للمعلومات اللي فوق
                result_df_display.to_excel(writer, index=False, sheet_name='الشجرة', startrow=4)
                
                workbook = writer.book
                worksheet = writer.sheets['الشجرة']
                worksheet.sheet_view.rightToLeft = True
                
                from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
                
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center', vertical='center')
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # أزرق
                header_font = Font(color="FFFFFF", bold=True, size=11)
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # أصفر للإجمالي
                grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # رصاصي للغياب
                
                # 1. وضع البيانات الأساسية في أعلى الإكسيل زي الصورة
                meta_data = [("الكنترول", control_name), ("المقرر", course_name), ("التاريخ", exam_date)]
                for i, (label, val) in enumerate(meta_data, start=1):
                    worksheet[f'A{i}'] = label
                    worksheet[f'B{i}'] = val
                    worksheet[f'A{i}'].border = thin_border
                    worksheet[f'B{i}'].border = thin_border
                    worksheet[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center') # توسيط
                    worksheet[f'B{i}'].alignment = center_align
                    worksheet[f'A{i}'].font = Font(bold=True)
                    worksheet[f'B{i}'].font = Font(bold=True)

                # 2. تنسيق رأس الجدول (الأزرق) وإضافة الفلتر
                for col_num in range(1, 5):
                    cell = worksheet.cell(row=5, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                worksheet.auto_filter.ref = f"A5:D{worksheet.max_row-1}"

                # 3. تنسيق صفوف الجدول (الرصاصي للغياب، والحدود للكل)
                for r_idx in range(6, worksheet.max_row + 1):
                    is_total = (r_idx == worksheet.max_row)
                    attendance_val = worksheet.cell(row=r_idx, column=3).value
                    
                    for c_idx in range(1, 5):
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        cell.border = thin_border
                        cell.alignment = center_align
                        
                        if is_total:
                            cell.fill = yellow_fill
                            cell.font = Font(bold=True)
                        elif attendance_val == 0 or attendance_val == '0':
                            cell.fill = grey_fill

                # 4. دمج (Merge) خلايا صف الإجمالي
                last_row = worksheet.max_row
                worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=2)
                worksheet.cell(row=last_row, column=1).value = "الإجمالي"

                # 5. ضبط عرض الأعمدة لتناسب النصوص
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 35
                worksheet.column_dimensions['C'].width = 15
                worksheet.column_dimensions['D'].width = 30
            
            st.markdown("<div style='display: flex; justify-content: flex-end; width: 100%;'>", unsafe_allow_html=True)
            st.download_button("📥 تحميل ملف الإكسيل المنسق", output.getvalue(), "توزيع_الشجرة.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
            st.markdown("</div>", unsafe_allow_html=True)
            st.markdown("<p dir='rtl' style='text-align: right; color: gray; margin-top: 10px; font-size: 14px;'>💡 نصيحة: للطباعة كـ PDF، قم بفتح ملف الإكسيل الذي تم تحميله، ثم اضغط (Ctrl+P) واختر Save as PDF</p>", unsafe_allow_html=True)

    st.markdown("---")
    if st.button("تفريغ البيانات لرفع ملف جديد"):
        st.session_state.base_df = None
        st.rerun()
